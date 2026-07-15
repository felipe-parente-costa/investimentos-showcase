"""Parser for the B3 "Movimentação" xlsx export (Área do Investidor / former CEI).

Expected columns: Entrada/Saída, Data, Movimentação, Produto, Instituição,
Quantidade, Preço unitário, Valor da Operação.

Conventions:
- Quantity is positive except for `transfer` operations, where the sign
  encodes direction (Credito = +, Debito = -). Buy/sell direction lives in
  the operation itself.
- Rows that do not affect positions or cash relevant to the portfolio
  (stock lending, subscription rights, refunds) are skipped and reported.
- Unknown movement types are never guessed: they are skipped with a reason
  so the user can extend the mapping deliberately.
"""

import re
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import BinaryIO

import openpyxl

from app.models.enums import AssetClass, Operation
from app.parsers.base import (
    ParseError,
    ParseResult,
    ParsedTransaction,
    ParserWarning,
    SkippedRow,
)

EXPECTED_HEADER = [
    "Entrada/Saída",
    "Data",
    "Movimentação",
    "Produto",
    "Instituição",
    "Quantidade",
    "Preço unitário",
    "Valor da Operação",
]

# (direction, movement type) -> Operation. Keys are lowercased.
OPERATION_MAP: dict[tuple[str, str], Operation] = {
    ("credito", "transferência - liquidação"): Operation.buy,
    ("debito", "transferência - liquidação"): Operation.sell,
    ("credito", "compra"): Operation.buy,
    # CDB purchases show up as a credit named "COMPRA / VENDA".
    ("credito", "compra / venda"): Operation.buy,
    ("debito", "venda"): Operation.sell,
    ("credito", "resgate"): Operation.sell,
    ("debito", "resgate"): Operation.sell,
    ("debito", "resgate antecipado"): Operation.sell,
    # The fractional share is already removed from custody by "Fração em
    # Ativos" (a transfer-out); "Leilão de Fração" is only the cash credit
    # from auctioning it. Mapping it to a sell would debit the fraction a
    # second time, so it is recorded as income (no quantity effect).
    ("credito", "leilão de fração"): Operation.yield_,
    ("credito", "desdobro"): Operation.split,
    ("credito", "dividendo"): Operation.dividend,
    ("credito", "juros sobre capital próprio"): Operation.jcp,
    ("credito", "rendimento"): Operation.yield_,
    ("credito", "bonificação em ativos"): Operation.bonus,
    # Position updates without cash (ticker conversions, custody moves,
    # fractions sent to auction). Direction is encoded in the quantity sign.
    ("credito", "atualização"): Operation.transfer,
    ("debito", "atualização"): Operation.transfer,
    ("credito", "transferência"): Operation.transfer,
    ("debito", "transferência"): Operation.transfer,
    ("debito", "fração em ativos"): Operation.transfer,
    # CDB maturity is exported with price/total 0, so the cash received is
    # unknown; a transfer-out closes the position at cost without
    # fabricating a realized loss.
    ("debito", "vencimento"): Operation.transfer,
}

_TRANSFERIDO = (
    "income custody transfer: credit/debit pair nets to zero, the income "
    "itself is recorded by the plain row"
)

SKIP_TYPES: dict[str, str] = {
    "empréstimo": "stock lending (position is mirrored by Transferência rows)",
    "reembolso": "lending refund, no position effect",
    "dividendo - transferido": _TRANSFERIDO,
    "juros sobre capital próprio - transferido": _TRANSFERIDO,
    "rendimento - transferido": _TRANSFERIDO,
    "cessão de direitos": "subscription rights, out of scope",
    "cessão de direitos - solicitada": "subscription rights, out of scope",
    "direito de subscrição": "subscription rights, out of scope",
    "direitos de subscrição - não exercido": "subscription rights, out of scope",
}

# A B3 negotiable code: starts with a letter, 5-6 alphanumerics, ends in a
# digit, with an optional fractional-market trailing letter. Unlike the older
# `[A-Z]{4}\d{1,2}` form, the alpha part may contain a digit, so BDRs whose
# code is not four clean letters (e.g. Meta's M1TA34) are recognized instead
# of falling through to the fixed-income fallback.
TICKER_RE = re.compile(r"^([A-Z][A-Z0-9]{3,4}\d[A-Z]?)\s*-\s*(.+)$")

FII_NAME_HINTS = ("IMOB", "FII")
ETF_NAME_HINTS = ("ETF", "ISHARES", "ÍNDICE", "INDICE")


class CeiParseError(ParseError):
    pass


def parse_cei_xlsx(file: BinaryIO | bytes) -> ParseResult:
    if isinstance(file, bytes):
        file = BytesIO(file)
    workbook = openpyxl.load_workbook(file, data_only=True)
    if "Movimentação" in workbook.sheetnames:
        sheet = workbook["Movimentação"]
    else:
        sheet = workbook[workbook.sheetnames[0]]

    rows = sheet.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows, [])]
    if header[: len(EXPECTED_HEADER)] != EXPECTED_HEADER:
        raise CeiParseError(
            f"Unexpected header {header!r}; expected {EXPECTED_HEADER!r}. "
            "Is this a B3 Movimentação export?"
        )

    data_rows = list(rows)

    # Tickers that take part in the B3 stock-lending program (any "Empréstimo"
    # row). Used to tell apart the two faces of "Atualização (Credito)": for a
    # lent ticker it is the B3 re-crediting shares that never left custody (a
    # phantom credit, skipped below); for any other ticker it is a legitimate
    # ticker conversion (B3 renaming a security to a new symbol) that must be
    # kept. The signal is purely the *presence* of an "Empréstimo" row for the
    # ticker in the same statement — not the contract, the lot size, or a
    # matching quantity (re-credits routinely have no equal-sized lend leg).
    # This assumes lending and conversion never coexist on one ticker.
    lending_tickers = {
        _parse_product(str(r[3]))[0]
        for r in data_rows
        if r is not None and str(r[2]).strip().lower() == "empréstimo"
    }

    transactions: list[ParsedTransaction] = []
    skipped: list[SkippedRow] = []
    warnings: list[ParserWarning] = []
    for row_number, row in enumerate(data_rows, start=2):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        (
            direction_raw,
            date_raw,
            movement_raw,
            product_raw,
            institution_raw,
            qty_raw,
            price_raw,
            total_raw,
        ) = row[:8]

        direction = str(direction_raw).strip().lower()
        movement = str(movement_raw).strip()
        movement_key = movement.lower()

        if movement_key in SKIP_TYPES:
            skipped.append(SkippedRow(row_number, movement, SKIP_TYPES[movement_key]))
            continue

        operation = OPERATION_MAP.get((direction, movement_key))
        if operation is None:
            skipped.append(
                SkippedRow(row_number, movement, f"unmapped movement type ({direction_raw})")
            )
            continue

        ticker, asset_name, asset_class = _parse_product(str(product_raw))

        # A lent ticker's "Atualização (Credito)" is a lending re-credit, not a
        # real position change (see lending_tickers note above): drop it so the
        # position is not inflated by phantom shares. Conversions (same
        # movement on a non-lent ticker) fall through and stay transfers.
        if (
            direction == "credito"
            and movement_key == "atualização"
            and ticker in lending_tickers
        ):
            skipped.append(
                SkippedRow(
                    row_number,
                    movement,
                    "lending re-credit (ticker participates in Empréstimo); "
                    "shares never left custody",
                )
            )
            continue

        quantity = _parse_decimal(qty_raw)
        unit_price = _parse_decimal(price_raw)
        total_value = _parse_decimal(total_raw)

        # A "Transferência - Liquidação" with no price is not a trade: it is
        # the custody leg of stock lending (shares moving to/from the
        # borrower). Mapping it as buy/sell at price 0 fabricates realized
        # losses and dilutes the average on return; as a transfer, the
        # quantity moves while the average price is preserved (the engine's
        # dormant_average survives full lend-outs).
        if (
            movement_key == "transferência - liquidação"
            and unit_price == 0
            and total_value == 0
        ):
            operation = Operation.transfer
        if total_value == 0 and quantity and unit_price:
            total_value = quantity * unit_price
        if operation is Operation.transfer and direction == "debito":
            quantity = -quantity

        tx_date = _parse_date(date_raw)

        # A *priced* "Transferência - Liquidação" on a lent ticker is
        # ambiguous: a real settled trade and a stock-lending recall/return are
        # identical rows in this export (defeitos 2/3). The data does not allow
        # a deterministic split, so the parser does not reclassify — it keeps
        # the buy/sell and flags the leg for manual reconciliation. Only lent
        # tickers are flagged; on a non-lent ticker the leg is unambiguously a
        # trade.
        if (
            movement_key == "transferência - liquidação"
            and unit_price > 0
            and ticker in lending_tickers
        ):
            warnings.append(
                ParserWarning(
                    row=row_number,
                    ticker=ticker,
                    date=tx_date,
                    quantity=quantity,
                    message=(
                        f"priced 'Transferência - Liquidação' on lent ticker "
                        f"{ticker} ({quantity} @ {unit_price} on {tx_date}): may "
                        "be a stock-lending return, not a trade — reconcile "
                        "manually"
                    ),
                )
            )

        institution = str(institution_raw).strip() if institution_raw else None

        transactions.append(
            ParsedTransaction(
                row=row_number,
                date=tx_date,
                ticker=ticker,
                asset_name=asset_name,
                asset_class=asset_class,
                operation=operation,
                quantity=quantity,
                unit_price=unit_price,
                total_value=total_value,
                notes=f"B3: {movement} ({direction_raw})",
                institution=institution,
            )
        )

    # Keep a ticker's asset class consistent across the file: B3 sometimes
    # relabels the same security over time (e.g. a FII listed first under a
    # fund name without an "IMOB/FII" hint, later with it), which would split
    # one ticker between stock and fii. A FII signal anywhere wins.
    fii_tickers = {
        t.ticker for t in transactions if t.asset_class is AssetClass.fii
    }
    for t in transactions:
        if t.ticker in fii_tickers and t.asset_class is not AssetClass.fii:
            t.asset_class = AssetClass.fii

    # The export lists rows newest-first (also within a day); reverse to
    # chronological order so intra-day sequencing is correct downstream.
    transactions.reverse()
    return ParseResult(transactions=transactions, skipped=skipped, warnings=warnings)


def _parse_product(product: str) -> tuple[str, str | None, AssetClass]:
    product = product.strip()
    match = TICKER_RE.match(product)
    if not match:
        # No "TICK11 - Name" prefix: Tesouro Direto and other fixed income.
        return product, None, AssetClass.fixed_income
    ticker, name = match.group(1), match.group(2).strip()
    upper = name.upper()
    if any(hint in upper for hint in FII_NAME_HINTS):
        asset_class = AssetClass.fii
    elif any(hint in upper for hint in ETF_NAME_HINTS):
        asset_class = AssetClass.etf
    else:
        asset_class = AssetClass.stock
    return ticker, name, asset_class


def _parse_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return datetime.strptime(str(value).strip(), "%d/%m/%Y").date()


def _parse_decimal(value: object) -> Decimal:
    if value is None:
        return Decimal("0")
    if isinstance(value, str):
        value = value.strip()
        if value in ("-", ""):
            return Decimal("0")
        value = value.replace(".", "").replace(",", ".") if "," in value else value
    return Decimal(str(value))
