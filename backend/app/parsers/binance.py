"""Parsers for the two Binance spot xlsx exports.

Binance offers two different spot exports; this module reads both and a
dispatcher picks the right one by inspecting the header row.

"Spot Trade History" (`parse_binance_spot_xlsx`):
- One row per execution (fill), with a Fee column.
- Header: Time, Pair, Side, Price, Executed, Amount, Fee.

"Spot Order History" (`parse_binance_order_xlsx`):
- One row per order, aggregating its fills, with a Status column
  (FILLED/CANCELED) and no Fee column. It has both an "Order Price"
  (the limit) and an "Average Price" (the realized execution price); the
  latter is the one comparable to Trade History's "Price", so it is what
  we store as unit_price (keeps import hashes aligned across exports).
- Header: Time, OrderNo, Pair, Type, Side, Order Price, Order Amount,
  Time, Executed, Average Price, Trading total, Status. Some labels carry
  a trailing footnote superscript (Type¹, Executed², Trading total³).

Shared layout quirks of the real files:
- Data sits in a sparse grid: a decorative preamble (name, e-mail, period),
  then a header row in non-adjacent columns, located by scanning.
- Executed/Amount/total are strings with the asset glued to the number
  ("0.00123BTC", "399.87177BRL", "34.8768009USDT"); base and quote assets
  are derived from those suffixes, not by splitting the pair symbol.
- Timestamps are in the export's local offset, declared in the
  "Period(UTC--3)" cell; they are converted to UTC and the UTC date is
  stored (original timestamp kept in notes).
"""

import re
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from io import BytesIO
from typing import BinaryIO, Callable

import openpyxl

from app.models.enums import AssetClass, Custody, Operation
from app.parsers.base import ParseError, ParseResult, ParsedTransaction, SkippedRow

HEADER_FIELDS = ["Time", "Pair", "Side", "Price", "Executed", "Amount", "Fee"]
ORDER_HEADER_FIELDS = [
    "Time",
    "Pair",
    "Side",
    "Executed",
    "Average Price",
    "Trading total",
    "Status",
]

SIDE_MAP = {"BUY": Operation.buy, "SELL": Operation.sell}

AMOUNT_RE = re.compile(r"^([0-9]+(?:\.[0-9]+)?)([A-Z0-9]+)$")
OFFSET_RE = re.compile(r"UTC([+-]\d+)")
_SUPERSCRIPTS = {ord(c): None for c in "¹²³⁴⁵"}

# Stablecoin we deliberately do not track as a wallet asset: BRL→USDT legs
# (USDTBRL) are not positions, and USDT-quoted crypto buys have their cost
# converted to BRL at import time instead.
UNTRACKED_BASE_ASSETS = {"USDT"}


class BinanceParseError(ParseError):
    pass


def parse_binance_spot_xlsx(file: BinaryIO | bytes) -> ParseResult:
    if isinstance(file, bytes):
        file = BytesIO(file)
    workbook = openpyxl.load_workbook(file, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    return _parse_trade_rows(rows)


def _parse_trade_rows(rows: list[tuple]) -> ParseResult:
    tz = _find_export_timezone(rows)
    header_index, columns = _find_header(rows)

    transactions: list[ParsedTransaction] = []
    skipped: list[SkippedRow] = []
    for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        values = {field: row[col] for field, col in columns.items()}
        if all(v is None or str(v).strip() == "" for v in values.values()):
            continue

        side = str(values["Side"]).strip().upper()
        operation = SIDE_MAP.get(side)
        if operation is None:
            skipped.append(SkippedRow(row_number, side, "unmapped trade side"))
            continue

        executed_qty, base_asset = _split_amount(values["Executed"], row_number)
        total_value, quote_asset = _split_amount(values["Amount"], row_number)
        fee_qty, fee_asset = _split_amount(values["Fee"], row_number)
        price = Decimal(str(values["Price"]).strip().replace(",", ""))

        if base_asset in UNTRACKED_BASE_ASSETS:
            # USDTBRL: buying the intermediate stablecoin, not a wallet asset.
            # Same rule as the Order History path — see UNTRACKED_BASE_ASSETS.
            skipped.append(
                SkippedRow(row_number, str(values["Pair"]), f"{base_asset} not tracked")
            )
            continue

        notes = f"Binance spot {values['Pair']} {side} at {values['Time']} (UTC{tz.utcoffset(None).total_seconds() / 3600:+.0f})"
        quantity = executed_qty
        if fee_asset == quote_asset:
            # Paid on top, in the quote currency: quantity received is
            # unaffected, and this is a real extra charge added to cost.
            fees = fee_qty
        elif fee_asset == base_asset:
            # Deducted from the base asset itself before it lands in the
            # wallet: quantity is net of the fee (what you actually hold),
            # and `fees` (fee converted to quote at the trade price) keeps
            # quantity*unit_price + fees == total_value (Amount) exact — the
            # real BRL spent is neither double-counted nor understated.
            quantity = executed_qty - fee_qty
            fees = fee_qty * price
        else:
            # Fee paid in a third asset (e.g. a BNB discount) comes from an
            # untracked balance: it neither reduces the base asset received
            # nor adds to the quote-currency cost recorded here.
            fees = Decimal("0")
            notes += f"; fee {fee_qty}{fee_asset} not converted"

        transactions.append(
            ParsedTransaction(
                row=row_number,
                date=_to_utc_date(values["Time"], tz),
                ticker=base_asset,
                asset_name=None,
                asset_class=AssetClass.crypto,
                operation=operation,
                quantity=quantity,
                unit_price=price,
                total_value=total_value,
                notes=notes,
                currency=quote_asset,
                fees=fees,
                institution="Binance",
                custody=Custody.binance,
            )
        )

    # The export lists trades newest-first; reverse to chronological order.
    transactions.reverse()
    return ParseResult(transactions=transactions, skipped=skipped)


def parse_binance_xlsx(file: BinaryIO | bytes) -> ParseResult:
    """Dispatch to the Trade or Order History parser by header shape.

    A Trade History has a Fee column; an Order History has a Status column.
    USDT-quoted rows from an Order History keep their USDT currency here and
    must be converted to BRL by `convert_usdt_to_brl` (which needs FX) before
    import.
    """
    if isinstance(file, bytes):
        file = BytesIO(file)
    workbook = openpyxl.load_workbook(file, data_only=True)
    rows = list(workbook[workbook.sheetnames[0]].iter_rows(values_only=True))
    if _has_header(rows, ORDER_HEADER_FIELDS):
        return _parse_order_rows(rows)
    if _has_header(rows, HEADER_FIELDS):
        return _parse_trade_rows(rows)
    raise BinanceParseError(
        "No Binance Spot Trade History (Fee column) or Order History "
        "(Status column) header found. Is this a Binance spot export?"
    )


def parse_binance_order_xlsx(file: BinaryIO | bytes) -> ParseResult:
    if isinstance(file, bytes):
        file = BytesIO(file)
    workbook = openpyxl.load_workbook(file, data_only=True)
    rows = list(workbook[workbook.sheetnames[0]].iter_rows(values_only=True))
    return _parse_order_rows(rows)


def _parse_order_rows(rows: list[tuple]) -> ParseResult:
    tz = _find_export_timezone(rows)
    header_index, columns = _find_order_header(rows)

    transactions: list[ParsedTransaction] = []
    skipped: list[SkippedRow] = []
    for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        values = {field: row[col] for field, col in columns.items()}
        if all(v is None or str(v).strip() == "" for v in values.values()):
            continue

        status = str(values["Status"]).strip().upper()
        if status != "FILLED":
            # CANCELED/EXPIRED/PARTIALLY_FILLED-then-canceled moved nothing.
            skipped.append(SkippedRow(row_number, status, "order not filled"))
            continue

        side = str(values["Side"]).strip().upper()
        operation = SIDE_MAP.get(side)
        if operation is None:
            skipped.append(SkippedRow(row_number, side, "unmapped trade side"))
            continue

        executed_qty, base_asset = _split_amount(values["Executed"], row_number)
        total_value, quote_asset = _split_amount(values["Trading total"], row_number)
        # "Average Price" is the realized execution price (matches Trade
        # History's "Price"); "Order Price" is only the limit and is ignored.
        price = Decimal(str(values["Average Price"]).strip().replace(",", ""))

        if base_asset in UNTRACKED_BASE_ASSETS:
            # USDTBRL: buying the intermediate stablecoin, not a wallet asset.
            skipped.append(
                SkippedRow(row_number, str(values["Pair"]), f"{base_asset} not tracked")
            )
            continue

        notes = (
            f"Binance order {values['Pair']} {side} at {values['Time']} "
            f"(UTC{tz.utcoffset(None).total_seconds() / 3600:+.0f})"
        )
        transactions.append(
            ParsedTransaction(
                row=row_number,
                date=_to_utc_date(values["Time"], tz),
                ticker=base_asset,
                asset_name=None,
                asset_class=AssetClass.crypto,
                operation=operation,
                quantity=executed_qty,
                unit_price=price,
                total_value=total_value,
                # Order History carries no fee column; fills' tiny taker fees
                # (~0.1%) are not recoverable here and are recorded as 0.
                fees=Decimal("0"),
                notes=notes,
                currency=quote_asset,  # BRL or USDT (converted later)
                institution="Binance",
                custody=Custody.binance,
            )
        )

    transactions.reverse()
    return ParseResult(transactions=transactions, skipped=skipped)


def convert_usdt_to_brl(
    transactions: list[ParsedTransaction], fx: Callable[[date], Decimal]
) -> list[ParsedTransaction]:
    """Rewrite USDT-quoted crypto buys to BRL using the order day's rate.

    USDT is treated as ~1 USD, so `fx` is the PTAX USD/BRL closing rate for
    the (UTC) order date. unit_price, total_value and fees are scaled and the
    currency becomes BRL, making every crypto cost basis BRL-denominated and
    consistent with the BRL-quoted rows and the BTC/ETH live quotes.
    """
    for tx in transactions:
        if tx.currency == "USDT":
            rate = fx(tx.date)
            tx.unit_price = tx.unit_price * rate
            tx.total_value = tx.total_value * rate
            tx.fees = tx.fees * rate
            tx.currency = "BRL"
            tx.notes += f"; USDT→BRL @ {rate} (PTAX)"
    return transactions


def _header_columns(row: tuple) -> dict[str, int]:
    # Strip footnote superscripts (Executed², Trading total³) and keep the
    # first column for repeated labels (Order History lists "Time" twice).
    cells: dict[str, int] = {}
    for col, c in enumerate(row):
        if c is None:
            continue
        key = str(c).translate(_SUPERSCRIPTS).strip()
        cells.setdefault(key, col)
    return cells


def _has_header(rows: list[tuple], fields: list[str]) -> bool:
    return any(
        all(f in _header_columns(row) for f in fields) for row in rows
    )


def _find_header(rows: list[tuple]) -> tuple[int, dict[str, int]]:
    for index, row in enumerate(rows):
        cells = _header_columns(row)
        if all(field in cells for field in HEADER_FIELDS):
            return index, {field: cells[field] for field in HEADER_FIELDS}
    raise BinanceParseError(
        "Header row with Time/Pair/Side/Price/Executed/Amount/Fee not found. "
        "Is this a Binance Spot Trade History export?"
    )


def _find_order_header(rows: list[tuple]) -> tuple[int, dict[str, int]]:
    for index, row in enumerate(rows):
        cells = _header_columns(row)
        if all(field in cells for field in ORDER_HEADER_FIELDS):
            return index, {field: cells[field] for field in ORDER_HEADER_FIELDS}
    raise BinanceParseError(
        "Header row with Time/Pair/Side/Executed/Average Price/Trading total/"
        "Status not found. Is this a Binance Spot Order History export?"
    )


def _find_export_timezone(rows: list[tuple]) -> timezone:
    for row in rows:
        for cell in row:
            if isinstance(cell, str) and "UTC" in cell and "Period" in cell:
                match = OFFSET_RE.search(cell.replace("--", "-"))
                if match:
                    return timezone(timedelta(hours=int(match.group(1))))
    # No period header: assume timestamps are already UTC.
    return timezone.utc


def _to_utc_date(value: object, tz: timezone) -> date:
    if isinstance(value, datetime):
        local = value
    else:
        local = datetime.strptime(str(value).strip(), "%Y-%m-%d %H:%M:%S")
    return local.replace(tzinfo=tz).astimezone(timezone.utc).date()


def _split_amount(value: object, row_number: int) -> tuple[Decimal, str]:
    text = str(value).strip().replace(",", "")
    match = AMOUNT_RE.match(text)
    if not match:
        raise BinanceParseError(f"Cannot parse amount {value!r} at row {row_number}")
    return Decimal(match.group(1)), match.group(2)
