"""B3 stock-lending reconciliation (Frente B — implemented 2026-07-08, after
the original 3 project phases; not part of the Fase 3 scope in docs/projeto.md).

The B3 Movimentação export writes the custody legs of the securities-lending
program (BTC) as rows indistinguishable from real trades: priced
"Transferência - Liquidação" legs, zero-priced custody legs, and phantom
"Atualização" re-credits. The discriminant lives in a separate filtered
export (filter "Outros"): `Empréstimo` events carrying ticker/date/qty.

This module:

- parses that events export (`parse_lending_events_xlsx`);
- collapses raw events into *contracts* (`collapse_contracts`): a contract
  registration and its fee line arrive as a pair of rows days apart, and
  renewals repeat monthly — grouping same (ticker, qty) events within
  ``PAIR_WINDOW_DAYS`` into one contract removes the date ambiguity that a
  raw ±3-day match suffers from (tested against the real fixtures, not
  assumed);
- reconciles a parsed Movimentação (`reconcile`): matched priced legs and
  zero-priced legs on lending tickers are custody, not trades — they are
  dropped, so the replayed position is real trades + corporate events only.
  Each contract claims at most one leg per direction (nearest first); a
  matched leg without a free slot stays a trade and carries a warning for
  manual confirmation. Legs the user already adjudicated on a brokerage
  note are passed in ``confirmed_trades`` and never touched (a genuine
  trade right next to a same-quantity contract event is undecidable from
  the files alone);
- emits truncation pairs against the brokerage gabarito (`truncation_legs`):
  project policy fixes the target PM at the brokerage's per-operation 2-decimal
  truncation, so after the full-precision replay each gabarito ticker gets
  a +1/-1 cost pair (fixed-point iterated) closing the gap. The engine
  stays full-precision Decimal; truncation enters as auditable data legs.
"""

from dataclasses import dataclass, field
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace
from typing import BinaryIO, FrozenSet

import openpyxl

from app.models.enums import AssetClass, Market, Operation
from app.parsers.base import ParseResult, ParsedTransaction, SkippedRow
from app.parsers.cei import (
    CeiParseError,
    EXPECTED_HEADER,
    _parse_date,
    _parse_decimal,
    _parse_product,
)
from app.services.portfolio import compute_positions

ZERO = Decimal("0")
# A leg matches a contract when it falls within this many days of the
# contract's [start, end] range — the same ±3 window validated empirically
# on the 6 corrected papers.
MATCH_WINDOW_DAYS = 3
# Registration + fee-line pairs of one contract arrive within days of each
# other; same (ticker, qty) events closer than this collapse into one
# contract. Monthly renewals (~28 days apart) stay separate.
PAIR_WINDOW_DAYS = 5
# Convergence threshold for the truncation pair: half a display cent.
HALF_CENT = Decimal("0.005")
Q10 = Decimal("0.0000000001")  # Numeric(24, 10) resolution

# User-adjudicated REAL trades that the event matching would classify as
# custody (a genuine buy/sell right next to a same-quantity contract event is
# undecidable from the files alone — only the brokerage note settles it).
# Same explicit-allowlist pattern as ASSET_CLASS_OVERRIDES: extend
# deliberately, never by heuristic. Entries are (ticker, date, quantity).
CONFIRMED_TRADES: FrozenSet[tuple[str, date, Decimal]] = frozenset()


@dataclass(frozen=True)
class RawLendingEvent:
    ticker: str
    date: date
    quantity: Decimal
    direction: str  # "credito" | "debito"


@dataclass(frozen=True)
class ReferenceEvent:
    """A reference-data row for the lending_events table: an `Empréstimo`
    contract event (qty > 0) or an `Atualização` credit. Not a transaction."""

    ticker: str
    date: date
    quantity: Decimal
    direction: str
    kind: str  # "emprestimo" | "atualizacao"


@dataclass
class LendingExportParse:
    """Categorized contents of a filtered Movimentação export (B3 filter
    "Outros" or "Reembolso e Empréstimos" — both share the header)."""

    events: list[ReferenceEvent] = None  # type: ignore[assignment]
    reembolsos: list[ParsedTransaction] = None  # type: ignore[assignment]
    skipped: list[SkippedRow] = None  # type: ignore[assignment]

    def __post_init__(self) -> None:
        self.events = self.events or []
        self.reembolsos = self.reembolsos or []
        self.skipped = self.skipped or []


@dataclass(frozen=True)
class Contract:
    """A lending contract event: one or more same-(ticker, qty) rows within
    PAIR_WINDOW_DAYS (registration + fee line, or an explicit return pair),
    represented by their date range."""

    ticker: str
    quantity: Decimal
    start: date
    end: date

    def distance(self, day: date) -> int:
        if self.start <= day <= self.end:
            return 0
        return min(abs((day - self.start).days), abs((day - self.end).days))


def parse_lending_export_xlsx(file: BinaryIO | bytes) -> LendingExportParse:
    """Single parser for the filtered Movimentação exports.

    Gotcha (also shown as help text on the import card): B3's "Empréstimos"
    filter does NOT return `Empréstimo`-type rows — the events come under
    the "Outros" filter; "Reembolso e Empréstimos" only carries the
    reimbursements.

    Categorization (design doc §3.8, approved 2026-07-08):
    - `Empréstimo` with qty > 0 and `Atualização` -> reference events
      (lending_events table); `Empréstimo` fee lines (qty 0) are redundant
      to the contract collapse and skipped;
    - `Reembolso` -> a normal `yield_` income transaction (real money:
      dividends passed through by the borrower); same hash formula as the
      main import, so a future overlap deduplicates itself;
    - everything else (corporate events, subscriptions) is skipped with a
      reason: those rows enter via the full Movimentação import.
    """
    if isinstance(file, bytes):
        file = BytesIO(file)
    try:
        workbook = openpyxl.load_workbook(file, data_only=True)
    except Exception as exc:  # zipfile.BadZipFile etc.: not an xlsx at all
        raise CeiParseError(f"Arquivo não é um xlsx válido: {exc}") from exc
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows, [])]
    if header[: len(EXPECTED_HEADER)] != EXPECTED_HEADER:
        raise CeiParseError(
            f"Unexpected header {header!r}; expected {EXPECTED_HEADER!r}. "
            "Is this a B3 Movimentação (filtro Outros / Reembolso e "
            "Empréstimos) export?"
        )
    parse = LendingExportParse()
    for row_number, row in enumerate(rows, start=2):
        if row is None or row[2] is None:
            continue
        movement = str(row[2]).strip()
        key = movement.lower()
        direction = str(row[0]).strip().lower()
        if key == "empréstimo":
            quantity = _parse_decimal(row[5])
            if quantity <= 0:
                parse.skipped.append(
                    SkippedRow(
                        row_number,
                        movement,
                        "lending fee/rent line (qty 0): redundant to the "
                        "contract collapse",
                    )
                )
                continue
            ticker, _, _ = _parse_product(str(row[3]))
            parse.events.append(
                ReferenceEvent(ticker, _parse_date(row[1]), quantity, direction, "emprestimo")
            )
        elif key == "atualização":
            ticker, _, _ = _parse_product(str(row[3]))
            parse.events.append(
                ReferenceEvent(
                    ticker, _parse_date(row[1]), _parse_decimal(row[5]), direction, "atualizacao"
                )
            )
        elif key == "reembolso":
            ticker, asset_name, asset_class = _parse_product(str(row[3]))
            parse.reembolsos.append(
                ParsedTransaction(
                    row=row_number,
                    date=_parse_date(row[1]),
                    ticker=ticker,
                    asset_name=asset_name,
                    asset_class=asset_class,
                    operation=Operation.yield_,
                    quantity=ZERO,
                    unit_price=ZERO,
                    total_value=_parse_decimal(row[7]),
                    notes=f"B3: Reembolso ({str(row[0]).strip()}) — repasse de "
                    "proventos de ativo emprestado",
                    institution=str(row[4]).strip() if row[4] else None,
                )
            )
        else:
            parse.skipped.append(
                SkippedRow(
                    row_number,
                    movement,
                    "não é evento de empréstimo; importe pela Movimentação completa",
                )
            )
    return parse


def parse_lending_events_xlsx(file: BinaryIO | bytes) -> list[RawLendingEvent]:
    """`Empréstimo` events only — the reconciler's projection of
    parse_lending_export_xlsx (one parser, two views)."""
    parse = parse_lending_export_xlsx(file)
    return [
        RawLendingEvent(e.ticker, e.date, e.quantity, e.direction)
        for e in parse.events
        if e.kind == "emprestimo"
    ]


def collapse_contracts(
    events: list[RawLendingEvent],
) -> dict[str, list[Contract]]:
    """Group same-(ticker, qty) events within PAIR_WINDOW_DAYS into one
    contract with a date range. Returns contracts per ticker, sorted."""
    dates: dict[tuple[str, Decimal], list[date]] = {}
    for event in events:
        dates.setdefault((event.ticker, event.quantity), []).append(event.date)

    contracts: dict[str, list[Contract]] = {}
    for (ticker, quantity), day_list in dates.items():
        day_list.sort()
        start = end = day_list[0]
        for day in day_list[1:]:
            if (day - end).days <= PAIR_WINDOW_DAYS:
                end = day
            else:
                contracts.setdefault(ticker, []).append(
                    Contract(ticker, quantity, start, end)
                )
                start = end = day
        contracts.setdefault(ticker, []).append(Contract(ticker, quantity, start, end))
    for ticker in contracts:
        contracts[ticker].sort(key=lambda c: (c.start, c.quantity))
    return contracts


def raw_match_candidates(
    events: list[RawLendingEvent],
    ticker: str,
    day: date,
    quantity: Decimal,
    window: int = MATCH_WINDOW_DAYS,
) -> list[date]:
    """Candidate RAW event dates for a leg (pre-collapse). Exists so the
    ambiguity-elimination claim is testable: with raw events a leg can match
    the registration and the fee line of one contract (two dates); after
    collapse_contracts it must match exactly one contract."""
    return sorted(
        {
            e.date
            for e in events
            if e.ticker == ticker
            and e.quantity == quantity
            and abs((e.date - day).days) <= window
        }
    )


@dataclass
class ReconcileStats:
    dropped_priced: int = 0
    dropped_zero_legs: int = 0
    confirmed_trades_kept: int = 0
    slot_conflicts_kept_as_trade: int = 0
    unmatched_after_horizon: int = 0


def reconcile(
    result: ParseResult,
    contracts: dict[str, list[Contract]],
    confirmed_trades: FrozenSet[tuple[str, date, Decimal]] = frozenset(),
) -> tuple[ParseResult, ReconcileStats]:
    """Classify a parsed Movimentação against lending contracts.

    Returns a new ParseResult where custody legs became SkippedRows and the
    parser's blanket lending warnings are replaced by the reconciler's
    verdicts: warnings remain only for slot conflicts (matched leg kept as
    trade pending manual confirmation) and for legs dated beyond the events
    file's horizon (unclassifiable without a fresh export).
    """
    stats = ReconcileStats()
    horizon = max(
        (c.end for lst in contracts.values() for c in lst), default=date.min
    )

    # Global greedy claim: nearest (distance, leg date) pair wins; one leg
    # per direction per contract. Deterministic by construction.
    legs = []
    for index, tx in enumerate(result.transactions):
        if (
            tx.ticker in contracts
            and "Transferência - Liquidação" in tx.notes
            and tx.operation in (Operation.buy, Operation.sell)
            and tx.unit_price > 0
        ):
            if (tx.ticker, tx.date, abs(tx.quantity)) in confirmed_trades:
                stats.confirmed_trades_kept += 1
                continue
            legs.append((index, tx))

    candidates = []
    for index, tx in legs:
        for contract in contracts[tx.ticker]:
            if contract.quantity != abs(tx.quantity):
                continue
            dist = contract.distance(tx.date)
            if dist <= MATCH_WINDOW_DAYS:
                candidates.append((dist, tx.date, contract.start, index, contract))
    candidates.sort(key=lambda c: (c[0], c[1], c[2]))

    claimed_slots: set[tuple[int, Operation]] = set()  # (id(contract), direction)
    claimed_legs: dict[int, Contract] = {}
    conflicted: set[int] = set()
    for dist, _, _, index, contract in candidates:
        if index in claimed_legs or index in conflicted:
            continue
        tx = result.transactions[index]
        slot = (id(contract), tx.operation)
        if slot in claimed_slots:
            conflicted.add(index)  # provisional; may claim another contract later
            continue
        claimed_slots.add(slot)
        claimed_legs[index] = contract
    conflicted -= set(claimed_legs)

    transactions: list[ParsedTransaction] = []
    skipped = list(result.skipped)
    resolved: set[tuple[str, date, Decimal]] = set()
    kept_with_warning: set[tuple[str, date, Decimal]] = set()

    for index, tx in enumerate(result.transactions):
        if index in claimed_legs:
            contract = claimed_legs[index]
            skipped.append(
                SkippedRow(
                    tx.row,
                    "Transferência - Liquidação",
                    f"lending custody leg: {tx.operation.value} {abs(tx.quantity)} "
                    f"@ {tx.unit_price} matches contract {contract.quantity} "
                    f"[{contract.start}..{contract.end}]; shares never changed owner",
                )
            )
            resolved.add((tx.ticker, tx.date, abs(tx.quantity)))
            stats.dropped_priced += 1
            continue
        if (
            tx.ticker in contracts
            and tx.operation is Operation.transfer
            and "Transferência - Liquidação" in tx.notes
            and tx.unit_price == 0
        ):
            skipped.append(
                SkippedRow(
                    tx.row,
                    "Transferência - Liquidação",
                    f"lending custody zero-leg ({tx.quantity:+}); shares never "
                    "changed owner",
                )
            )
            resolved.add((tx.ticker, tx.date, abs(tx.quantity)))
            stats.dropped_zero_legs += 1
            continue
        if index in conflicted:
            kept_with_warning.add((tx.ticker, tx.date, abs(tx.quantity)))
            stats.slot_conflicts_kept_as_trade += 1
        transactions.append(tx)

    # Rebuild warnings from the reconciler's point of view.
    warnings = []
    for warning in result.warnings:
        key = (warning.ticker, warning.date, abs(warning.quantity))
        if key in resolved:
            continue  # classified as custody: warning resolved
        if key in kept_with_warning:
            warning.message += (
                " [reconciler: matched a contract whose slot is taken; kept as "
                "trade — confirm on the brokerage note]"
            )
            warnings.append(warning)
            continue
        if key in confirmed_trades:
            continue  # user-adjudicated real trade
        if warning.date > horizon:
            warning.message += (
                " [reconciler: dated beyond the lending-events export horizon "
                f"({horizon}); export a fresh 'Outros' file to classify]"
            )
            warnings.append(warning)
            stats.unmatched_after_horizon += 1
            continue
        # No contract within the window and inside the horizon: a real trade.
    return (
        ParseResult(transactions=transactions, skipped=skipped, warnings=warnings),
        stats,
    )


def _adapt(tx: ParsedTransaction, market: Market, currency: str) -> SimpleNamespace:
    return SimpleNamespace(
        date=tx.date,
        ticker=tx.ticker,
        asset_name=tx.asset_name,
        asset_class=tx.asset_class,
        market=market,
        currency=tx.currency or currency,
        operation=tx.operation,
        quantity=tx.quantity,
        unit_price=tx.unit_price,
        fees=tx.fees,
        total_value=tx.total_value,
        custody=tx.custody,
        custody_from=None,
        custody_to=None,
        indexer=tx.indexer,
        institution=tx.institution,
    )


def truncation_legs(
    transactions: list[ParsedTransaction],
    gabarito: dict[str, tuple[Decimal, Decimal]],
    market: Market = Market.br,
    currency: str = "BRL",
) -> list[ParsedTransaction]:
    """+1/-1 cost pairs closing the gap between the full-precision replay and
    the brokerage-truncated gabarito {ticker: (qty, pm)}, per project PM policy.

    Fails loud if the replayed quantity does not match the gabarito —
    a truncation pair fixes rounding, never quantity.
    """
    legs: list[ParsedTransaction] = []
    for ticker, (target_qty, target_pm) in gabarito.items():
        subset = [t for t in transactions if t.ticker == ticker]
        if not subset:
            raise ValueError(f"{ticker}: no transactions to truncate against")
        target_cost = target_qty * target_pm
        # One day after the ticker's last row: sharing a date with an
        # average-price transfer round trip (the plain "Transferência"
        # account pairs) halves the pair's cost effect per round trip and
        # stalls convergence. A day of its own keeps the effect exact and
        # the pair is quantity-neutral, so no history artifact is created.
        last = max(t.date for t in subset) + timedelta(days=1)
        pair_price = ZERO
        pos = None
        for _ in range(20):
            pair = _pair(ticker, last, pair_price)
            replay = [_adapt(t, market, currency) for t in subset + pair]
            pos = compute_positions(replay).positions[(ticker, None)]
            if pos.quantity != target_qty:
                raise ValueError(
                    f"{ticker}: replay quantity {pos.quantity} != gabarito "
                    f"{target_qty}; truncation pair cannot fix quantity"
                )
            delta = pos.total_cost - target_cost
            if abs(delta) < HALF_CENT:
                legs.extend(pair)
                break
            pair_price = (pair_price + delta).quantize(Q10)
        else:
            raise ValueError(f"{ticker}: truncation pair did not converge")
    return legs


def store_reference_events(db, events: list[ReferenceEvent]) -> int:
    """Insert reference events not yet known (timeline-extension idempotency
    over the natural key). Returns how many were added; the rest were
    already in the table."""
    from sqlalchemy import select

    from app.models.lending_event import LendingEventRecord

    added = 0
    for event in events:
        exists = db.execute(
            select(LendingEventRecord.id).where(
                LendingEventRecord.ticker == event.ticker,
                LendingEventRecord.date == event.date,
                LendingEventRecord.quantity == event.quantity,
                LendingEventRecord.direction == event.direction,
                LendingEventRecord.kind == event.kind,
            )
        ).first()
        if exists:
            continue
        db.add(
            LendingEventRecord(
                ticker=event.ticker,
                date=event.date,
                quantity=event.quantity,
                direction=event.direction,
                kind=event.kind,
            )
        )
        added += 1
    db.commit()
    return added


def load_contracts(db) -> dict[str, list[Contract]]:
    """Contracts collapsed from the lending_events table — the production
    counterpart of parse-then-collapse over a file."""
    from sqlalchemy import select

    from app.models.lending_event import LendingEventRecord

    rows = (
        db.execute(
            select(LendingEventRecord).where(LendingEventRecord.kind == "emprestimo")
        )
        .scalars()
        .all()
    )
    events = [
        RawLendingEvent(r.ticker, r.date, r.quantity, r.direction) for r in rows
    ]
    return collapse_contracts(events)


def _pair(ticker: str, on: date, price: Decimal) -> list[ParsedTransaction]:
    if price == 0:
        return []
    plus_price = -price if price < 0 else ZERO
    minus_price = price if price > 0 else ZERO

    def leg(qty: Decimal, unit_price: Decimal, half: str) -> ParsedTransaction:
        return ParsedTransaction(
            row=0,
            date=on,
            ticker=ticker,
            asset_name=None,
            asset_class=AssetClass.stock,
            operation=Operation.buy,
            quantity=qty,
            unit_price=unit_price,
            total_value=qty * unit_price,
            notes=(
                f"[par-truncagem-corretora] {half}: ajusta o custo do replay em "
                "precisão cheia à convenção truncada da corretora (política de PM "
                "do projeto); qty líquida do par = 0"
            ),
        )

    return [leg(Decimal(1), plus_price, "1/2"), leg(Decimal(-1), minus_price, "2/2")]
